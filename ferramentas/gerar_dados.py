#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera dados/dados.json a partir das planilhas Grupo 1..4.xlsx.

Regras estruturais das planilhas (ver docs/prd.md §9):
  - 1 aba por linha, nomeada com o número da linha (string, zero à esquerda preservado).
  - Os 3 dias do evento ficam EMPILHADOS verticalmente dentro da aba.
  - Cada dia tem 1..N turnos; cada turno tem 1..N tabelas (TAB - 0101, 0102, ...)
    dispostas HORIZONTALMENTE, cada uma com 6 colunas:
    IDA | VOLTA | TÉRM. | T. VIAG. | LINHA | VEIC.
  - Linhas de continuação (VOLTA em branco) têm os valores deslocados uma coluna
    à direita e NÃO são saídas novas: o IDA delas é o TÉRMINO da viagem anterior.
  - O painel lateral "FREQUÊNCIA DA LINHA" é a fonte autoritativa das saídas:
    len(saidas) == totalViagens == "N VIAGENS".
  - Horários após a meia-noite chegam como datetime com ano 1900 e são
    normalizados para "24:MM", "25:MM"..., preservando a ordem cronológica.

Uso:  python ferramentas/gerar_dados.py
"""

import copy
import datetime
import json
import re
import sys
from collections import OrderedDict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
PLANILHAS = RAIZ / "Passo a paço"
SAIDA = RAIZ / "dados" / "dados.json"
RELATORIO = RAIZ / "ferramentas" / "relatorio_conversao.txt"

DIAS = ["2026-09-05", "2026-09-06", "2026-09-07"]
GRUPOS = ["Grupo 1.xlsx", "Grupo 2.xlsx", "Grupo 3.xlsx", "Grupo 4.xlsx"]

EPOCA = datetime.date(1899, 12, 31)

# ---------------------------------------------------------------- referências

# Ponto de embarque por linha — as 83 linhas da arte da pág. 5 da OS 057.
# Nem todas têm aba de escala; o script só publica as que têm.
PONTO_POR_LINHA = {
    "P1": ["004", "203", "205", "207", "209", "214", "223", "227", "403", "427", "452",
           "502", "507", "606", "608", "609", "610", "611", "612", "625", "704", "706",
           "708", "711", "713"],
    "P2": ["011", "101", "110", "111", "112", "113", "119", "120", "121", "123", "126",
           "127", "129", "211", "219", "221", "302", "305", "306", "320", "321", "324"],
    "P3": ["002", "515", "517", "535", "540", "600", "602", "604", "616", "619", "621",
           "650", "651", "652", "676", "690"],
    "P4": ["315", "319", "330", "340", "356", "357", "422", "439", "440", "442", "443",
           "444", "447", "448", "454", "455", "457", "461", "500", "560", "640"],
}

ENTRADA = RAIZ / "ferramentas" / "entrada"
ARQ_NOMES = ENTRADA / "nomes_linhas.json"

# A escala aparece solta numa célula ("DOMINGO") ou enterrada num título
# ("PROPOSTA QUADRO OPERACIONAL, SÁBADO E DOMINGO, LINHA 123"). As alternativas
# vão da mais longa para a mais curta, senão "SÁBADO E DOMINGO" casaria só como
# "SÁBADO".
ESCALA_NO_TEXTO = re.compile(
    r"SEGUNDA A DOMINGO|S[ÁA]BADO E DOMINGO|DOMINGO|S[ÁA]BADO")

# Quando a mesma linha aparece em mais de um arquivo (aconteceu com a 219, que
# ficou nos Grupos 1 e 2 com versões diferentes), vale a aba do grupo cujo número
# corresponde ao ponto de embarque da linha — nunca a ordem de leitura.
GRUPO_DO_PONTO = {"P1": "Grupo 1.xlsx", "P2": "Grupo 2.xlsx",
                  "P3": "Grupo 3.xlsx", "P4": "Grupo 4.xlsx"}

PONTOS = [
    {"id": "P1", "numero": 1, "nome": "Ponto de Ônibus 01", "cor": "verde",
     "corHex": "#3FA45B", "zonas": "Zonas Sul, Centro-Sul e Centro-Oeste",
     "hotspot": {"x": 25.0, "y": 28.4}},
    {"id": "P2", "numero": 2, "nome": "Ponto de Ônibus 02", "cor": "azul",
     "corHex": "#2B3990", "zonas": "Zona Oeste",
     "hotspot": {"x": 25.0, "y": 43.3}},
    {"id": "P3", "numero": 3, "nome": "Ponto de Ônibus 03", "cor": "laranja",
     "corHex": "#F5821F", "zonas": "Zona Leste",
     "hotspot": {"x": 30.3, "y": 63.2}},
    {"id": "P4", "numero": 4, "nome": "Ponto de Ônibus 04", "cor": "vermelho",
     "corHex": "#E31E24", "zonas": "Zona Norte",
     "hotspot": {"x": 36.7, "y": 72.5}},
]

# Os nomes/itinerários vêm de ferramentas/entrada/nomes_linhas.json, editado à
# mão (ver ferramentas/entrada/LEIA-ME.md). Ficam fora deste arquivo de
# propósito: são a parte que muda a cada edição da OS, e quem edita não precisa
# mexer em Python.
#
# NENHUMA aba de planilha traz itinerário — as abas só têm horário. Os nomes que
# já existem foram transcritos das tabelas de viagens extras da OS 057 (31
# linhas), e vêm abreviados de lá mesmo.
#
# O TIPO DE ESCALA (domingo, sábado, sábado e domingo) não é declarado em lugar
# nenhum além da própria aba: ler_escala_base o extrai do cabeçalho. Só entra no
# relatório — a tela nunca o exibe.

def ler_entrada(caminho):
    """Lê um JSON de ferramentas/entrada/, com falha explícita e legível."""
    if not caminho.exists():
        sys.exit("ERRO: arquivo de entrada ausente: %s. "
                 "Veja ferramentas/entrada/LEIA-ME.md." % caminho)
    try:
        with open(caminho, encoding="utf-8") as f:
            dados = json.load(f)
    except json.JSONDecodeError as e:
        sys.exit("ERRO: %s não é JSON válido (linha %d, coluna %d): %s"
                 % (caminho.name, e.lineno, e.colno, e.msg))
    if not isinstance(dados, dict):
        sys.exit("ERRO: %s deve ser um objeto {\"linha\": ...}" % caminho.name)
    return dados


def ler_nomes():
    nomes = ler_entrada(ARQ_NOMES)
    for num, valor in nomes.items():
        if not re.match(r"^\d{3}$", str(num)):
            sys.exit("ERRO: nomes_linhas.json: chave %r não é um número de linha "
                     "de 3 dígitos" % num)
        if valor is not None and not isinstance(valor, str):
            sys.exit("ERRO: nomes_linhas.json: %s deve ter texto ou null" % num)
    return {k: (v.strip() or None) if isinstance(v, str) else None
            for k, v in nomes.items()}


# Concessionárias citadas na OS 057. A célula da planilha mistura
# "EMPRESA - TIPO DE VEÍCULO" com apenas "TIPO DE VEÍCULO", então a empresa só é
# atribuída quando bate com esta lista — nunca por posição na string.
EMPRESAS = ["SÃO PEDRO", "INTEGRAÇÃO", "VIA VERDE", "COROADO", "GLOBAL", "LÍDER", "VEGA"]

TODAS = "todas"

AVISOS = [
    {"id": "corredor-bb-19h", "dias": DIAS, "linhas": TODAS, "severidade": "alta",
     "texto": "A partir das 19h o corredor de ônibus à direita do Terminal Central, "
              "próximo ao Banco do Brasil, fica inoperante. Usar apenas o corredor "
              "próximo à Igreja e Praça da Matriz."},
    {"id": "desvio-itamaraca", "dias": DIAS, "linhas": TODAS, "severidade": "media",
     "texto": "Linhas que trafegam na Rua Itamaracá desviam pela Rua da Instalação "
              "para acessar o Terminal Central."},
    {"id": "paradas-inoperantes-00h", "dias": DIAS, "linhas": TODAS, "severidade": "alta",
     "texto": "Após as 00h as paradas Dom Bosco e Colégio Militar ficam inoperantes; "
              "permanece ativo apenas o Terminal da Matriz."},
    {"id": "onibus-extras-00h", "dias": DIAS, "linhas": TODAS, "severidade": "info",
     "texto": "25 ônibus extras disponibilizados a partir das 00h para apoio à operação."},
    {"id": "frota-apoio-t1", "dias": DIAS, "linhas": TODAS, "severidade": "info",
     "texto": "Frota de apoio estacionada nas proximidades do T1: 15 articulados, "
              "3 trucados e 7 convencionais, reduzida gradativamente conforme demanda."},
    {"id": "orientacao-fiscais", "dias": DIAS, "linhas": TODAS, "severidade": "info",
     "texto": "Seguir as orientações dos Fiscais de Transportes e Fiscais da ACOP "
              "presentes no local."},
    {"id": "viagem-extra-00h", "dias": DIAS, "severidade": "alta",
     "linhas": ["101", "113", "120", "121", "126", "214", "216", "219", "227", "315",
                "321", "356", "422", "440", "443", "448", "500", "535", "540", "560",
                "604", "608", "612", "621", "640", "650", "652", "676", "705", "713"],
     "texto": "Realiza mais uma viagem a partir das 00h."},
    {"id": "opera-ate-centro", "dias": DIAS, "severidade": "alta",
     "linhas": ["011", "216", "302", "305", "320", "321", "357", "444", "560", "621"],
     "texto": "Opera até a área central da cidade durante os dias do evento."},
    {"id": "604-676-domingo-18h", "dias": ["2026-09-07"], "severidade": "alta",
     "linhas": ["604", "676"],
     "texto": "Programação especial no dia 07/09 a partir das 18h, para apoio ao evento."},
    {"id": "121-substitui-019", "dias": ["2026-09-07"], "severidade": "alta",
     "linhas": ["121"],
     "texto": "No dia 07/09 substitui a operação da linha 019, com frota de 2 veículos."},
    {"id": "terminal-302-305", "dias": DIAS, "severidade": "alta",
     "linhas": ["302", "305"],
     "texto": "De 01/09 a 09/09 o terminal das linhas 302 e 305 fica transferido "
              "para a Praça da Saudade."},
]

# ------------------------------------------------------------------ utilidades

avisos_conversao = []


def avisar(msg):
    avisos_conversao.append(msg)


def eh_hora(v):
    return isinstance(v, (datetime.datetime, datetime.time))


def hora(v):
    """Converte célula de horário para 'HH:MM'. Ano 1900 = madrugada => 24h+."""
    if isinstance(v, datetime.datetime):
        dias_extra = (v.date() - EPOCA).days
        return "%02d:%02d" % (v.hour + 24 * max(dias_extra, 0), v.minute)
    if isinstance(v, datetime.time):
        return "%02d:%02d" % (v.hour, v.minute)
    return None


def duracao(v):
    """Igual a hora(), mas para durações (T. VIAG., JORN.) — nunca soma 24h."""
    if isinstance(v, datetime.datetime):
        return "%02d:%02d" % (v.hour + 24 * max((v.date() - EPOCA).days, 0), v.minute)
    if isinstance(v, datetime.time):
        return "%02d:%02d" % (v.hour, v.minute)
    return None


def txt(v):
    return v.strip() if isinstance(v, str) and v.strip() else None


def eh_data_evento(v):
    if isinstance(v, datetime.datetime) and v.year == 2026 and v.month == 9:
        return v.day in (5, 6, 7)
    if isinstance(v, str) and re.match(r"\s*0[567]/09/2026", v):
        return True
    return False


def celula(linha, col):
    return linha[col] if 0 <= col < len(linha) else None


def em_minutos(h):
    return int(h[:2]) * 60 + int(h[3:]) if h else None


def de_minutos(m):
    return "%02d:%02d" % (m // 60, m % 60)


def normalizar_monotonico(sequencia, campos):
    """Empurra para +24h todo horário que 'anda para trás' na sequência.

    Necessário porque a madrugada nem sempre chega como serial >24h: correções
    digitadas à mão entram como hora simples (01:20 em vez de 25:20). Dentro de
    uma mesma tabela, os horários de um veículo só avançam — então recuar
    significa ter cruzado a meia-noite.
    """
    LIMITE = 12 * 60  # recuo maior que isso = virada de meia-noite, não sobreposição
    ultimo = -10 ** 6
    for item in sequencia:
        for campo in campos:
            h = item.get(campo)
            if not h:
                continue
            m = em_minutos(h)
            # Só empurra num recuo GRANDE. Recuos pequenos são sobreposições
            # reais da planilha (visto na 227: viagem parte 00:00 enquanto a
            # anterior termina 00:14) e empurrá-las geraria 48:00.
            while ultimo - m > LIMITE:
                m += 24 * 60
            item[campo] = de_minutos(m)
            ultimo = max(ultimo, m)


# --------------------------------------------------------------------- parsing

def achar_blocos_de_dia(linhas):
    inicios = []
    for i, r in enumerate(linhas):
        if any(eh_data_evento(c) for c in r[:4]):
            inicios.append(i)
    inicios = sorted(set(inicios))
    return [(ini, inicios[k + 1] if k + 1 < len(inicios) else len(linhas))
            for k, ini in enumerate(inicios)]


def ler_painel_frequencia(linhas, ini, fim, limite):
    """Retorna [(hora, intervalo)] na ordem: par de colunas por par, de cima p/ baixo.

    O painel pode ser MAIS ALTO que o bloco do dia (visto na aba 640, onde ele
    invade 2 linhas do bloco seguinte). Por isso a leitura vai além de `fim`,
    parando na primeira célula vazia da coluna ou em `limite` — o cabeçalho do
    painel do próximo dia.
    """
    cab = None
    pares = []
    for i in range(ini, fim):
        cols = [j for j, c in enumerate(linhas[i]) if txt(c) == "HORA"]
        if cols:
            cab = i
            pares = [(j, j + 1) for j in cols]
            break
    if cab is None:
        return []
    saida = []
    for ch, cf in pares:
        for i in range(cab + 1, limite):
            h = celula(linhas[i], ch)
            if not eh_hora(h):
                break
            f = celula(linhas[i], cf)
            # Na linha 305 a coluna FREQ. traz o terminal de origem
            # ("KM 41" / "CENTRO") em vez do intervalo.
            saida.append({
                "hora": hora(h),
                "intervalo": duracao(f) if eh_hora(f) else None,
                "origem": txt(f) if isinstance(f, str) else None,
            })
    normalizar_monotonico(saida, ["hora"])
    return saida


def rotulo(v):
    if isinstance(v, str) and v.strip():
        return v.strip()
    if isinstance(v, (int, float)):
        return ("%g" % v)
    return None


def ler_viagem(r, c):
    """Lê as 6 colunas de uma tabela.

    Três formatos aparecem nas planilhas:
      normal      IDA | VOLTA | TÉRM. | T. VIAG. | LINHA | VEIC.
      continuação IDA | (vazio) | TÉRM. | (vazio) | T. VIAG.   -> valores deslocados
                  Perna de retorno: o IDA repete o TÉRMINO anterior. NÃO é saída.
      baixada     (vazio) | VOLTA | TÉRM. | T. VIAG. | LINHA | VEIC.
                  A viagem começa no retorno, após intervalo no outro terminal.
                  É saída, mas a partida é o VOLTA. Só ocorre na linha 305.
    """
    c0, c1, c2 = celula(r, c), celula(r, c + 1), celula(r, c + 2)

    if eh_hora(c0) and eh_hora(c1):
        return {
            "ida": hora(c0), "volta": hora(c1), "termino": hora(c2),
            "tempoViagem": duracao(celula(r, c + 3)),
            "linha": rotulo(celula(r, c + 4)), "veiculo": rotulo(celula(r, c + 5)),
            "tipo": "normal", "continuacao": False,
        }

    if eh_hora(c0) and eh_hora(c2):
        return {
            "ida": hora(c0), "volta": None, "termino": hora(c2),
            "tempoViagem": duracao(celula(r, c + 4)),
            "linha": None, "veiculo": None,
            "tipo": "continuacao", "continuacao": True,
        }

    if c0 is None and eh_hora(c1):
        return {
            "ida": None, "volta": hora(c1), "termino": hora(c2),
            "tempoViagem": duracao(celula(r, c + 3)),
            "linha": rotulo(celula(r, c + 4)), "veiculo": rotulo(celula(r, c + 5)),
            "tipo": "baixada", "continuacao": False,
        }

    return None


def ler_turnos(bloco, aba, dia):
    """Extrai os turnos do bloco de um dia."""
    marcas = []
    for i, r in enumerate(bloco):
        for c in r:
            s = txt(c)
            if s and re.match(r"^\d+\s*[ºo°]?\s*TURNO", s, re.I):
                marcas.append((i, s))
                break
    if not marcas:
        avisar("%s/%s: nenhum turno encontrado" % (aba, dia))
        return [], None

    turnos = []
    empresa = None
    for k, (ini, nome) in enumerate(marcas):
        fim = marcas[k + 1][0] if k + 1 < len(marcas) else len(bloco)
        trecho = bloco[ini:fim]

        linha_tab = None
        for i, r in enumerate(trecho):
            cols = {j: txt(c) for j, c in enumerate(r)
                    if txt(c) and txt(c).upper().startswith("TAB")}
            if cols:
                linha_tab = (i, cols)
                break
        if not linha_tab:
            avisar("%s/%s/%s: sem linha de TAB" % (aba, dia, nome))
            continue
        i_tab, tabs = linha_tab

        i_cab = None
        for i in range(i_tab + 1, len(trecho)):
            if any(txt(c) == "IDA" for c in trecho[i]):
                i_cab = i
                break
        if i_cab is None:
            avisar("%s/%s/%s: sem cabeçalho IDA" % (aba, dia, nome))
            continue

        tabelas = OrderedDict(
            (col, {"tab": rot, "veiculo": None, "tipoVeiculo": None,
                   "jornada": None, "viagens": []})
            for col, rot in sorted(tabs.items())
        )

        for r in trecho[i_cab + 1:]:
            marcador = txt(celula(r, 2))
            if marcador and marcador.upper().startswith("JORN"):
                for col, tb in tabelas.items():
                    tb["jornada"] = duracao(celula(r, col))
                break
            for col, tb in tabelas.items():
                v = celula(r, col)
                if isinstance(v, str) and v.strip():
                    bruto = v.strip()
                    achou = next((e for e in EMPRESAS if bruto.upper().startswith(e)), None)
                    if achou:
                        empresa = empresa or achou
                        bruto = bruto[len(achou):].lstrip(" -").strip()
                    tb["tipoVeiculo"] = bruto or None
                    continue
                viagem = ler_viagem(r, col)
                if viagem:
                    if viagem["veiculo"]:
                        tb["veiculo"] = viagem["veiculo"]
                    tb["viagens"].append(viagem)

        for tb in tabelas.values():
            normalizar_monotonico(tb["viagens"], ["ida", "volta", "termino"])

        turnos.append({"nome": nome, "tabelas": list(tabelas.values())})

    return turnos, empresa


def ler_notas(linhas):
    """Texto livre que a planilha traz fora do quadro: observações operacionais,
    marcações de intervalo e a legenda dos códigos de viagem (ex. 305.1)."""
    observacoes = []
    legenda = []
    vistas = []

    def repetida(s):
        # A mesma observação é redigitada em cada bloco de dia, com variações de
        # pontuação e erros de digitação ("ATENDERÁ P CENTRO!" vs "O CENTRO").
        # Compara por similaridade e guarda só a primeira redação.
        k = re.sub(r"[^A-Z0-9]+", "", s.upper())
        if any(SequenceMatcher(None, k, v).ratio() > 0.92 for v in vistas):
            return True
        vistas.append(k)
        return False

    for r in linhas:
        textos = [txt(c) for c in r]
        for s in textos:
            if not s:
                continue
            alto = s.upper()
            if alto.startswith("OBS") or alto.startswith("INTERVALO"):
                limpo = re.sub(r"^OBS[:.\s-]*", "", s).strip()
                if limpo and not repetida(limpo):
                    observacoes.append(limpo)
        # legenda: <código> | <veíc> | <qtd> | <descrição>
        vals = [c for c in r if c is not None]
        if len(vals) == 4 and isinstance(vals[3], str):
            cod = rotulo(vals[0])
            if cod and re.match(r"^\d{3}(\.\d+)?$", cod):
                item = {"codigo": cod, "veiculos": rotulo(vals[1]),
                        "viagens": rotulo(vals[2]), "descricao": vals[3].strip()}
                if item not in legenda:
                    legenda.append(item)
    return observacoes, legenda


def ler_escala_base(linhas):
    """Metadados da aba que NAO e do evento: escala regular da linha.

    Essas abas tem um unico bloco, com data de referencia propria (de 2021 a
    2026) e tipo de dia generico (DOMINGO / SABADO E DOMINGO). O QUADRO delas e
    publicado normalmente (ler_aba replica o bloco nos 3 dias); estes metadados
    ficam so para o relatorio: a tela nunca exibe escala nem data.
    """
    ref = None
    candidatos = []
    for i, r in enumerate(linhas[:5], 1):
        for c in r:
            if i == 2:
                if isinstance(c, datetime.datetime) and c.year > 1900:
                    ref = ref or c.strftime("%Y-%m-%d")
                elif isinstance(c, str) and re.match(r"\d{2}/\d{2}/\d{4}", c.strip()):
                    d, m, a = c.strip()[:10].split("/")
                    ref = ref or "%s-%s-%s" % (a, m, d)
            s = txt(c)
            if not s:
                continue
            alto = re.sub(r"\s+", " ", s.upper())
            # "FREQUÊNCIA DA LINHA 112" é o título do painel lateral, não a
            # escala: em 5 abas (112, 507, 515, 609, 625) ele cai na mesma
            # linha e era capturado como se fosse o tipo de dia.
            # "SEM COBRADOR AOS DOMINGOS" fala do cobrador, não da escala.
            if "FREQU" in alto or "TURNO" in alto or "COBRADOR" in alto:
                continue
            achado = ESCALA_NO_TEXTO.search(alto)
            if achado:
                candidatos.append(achado.group(0))
    # A redação mais completa vence: "SÁBADO E DOMINGO" ganha de "DOMINGO"
    # quando os dois aparecem na mesma aba.
    tipo = max(candidatos, key=len) if candidatos else None
    return {"dataReferencia": ref, "tipoDia": tipo}


def ler_bloco(linhas, ini, fim, limite, aba, dia):
    """Le UM bloco de escala — o de um dia do evento ou a aba regular inteira.

    Mesmo formato de saida nos dois casos, para que a UI nao precise distinguir
    linha com e sem programacao especial.
    """
    bloco = linhas[ini:fim]

    freq = ler_painel_frequencia(linhas, ini, fim, limite)
    turnos, empresa = ler_turnos(bloco, aba, dia)

    total = None
    tipo_dia = None
    for r in bloco:
        for c in r:
            s = txt(c)
            if not s:
                continue
            if "VIAGENS" in s.upper():
                m = re.search(r"(\d+)", s)
                if m:
                    total = int(m.group(1))
            if "PASSO A PAÇO" in s.upper():
                tipo_dia = s

    # PASSAGENS pelo ponto de embarque = coluna VOLTA.
    # IDA e a saida do terminal de origem, no bairro. VOLTA e o momento em
    # que o veiculo esta no Centro, que e onde ficam os 4 pontos da Av.
    # Epaminondas. O operador no ponto precisa da VOLTA, nao da IDA.
    # O horario e APROXIMADO e a UI deve dizer isso.
    passagens = sorted(
        v["volta"] for t in turnos for tb in t["tabelas"] for v in tb["viagens"]
        if v["tipo"] != "continuacao" and v["volta"])

    # Viagens "cortadas": comecam no retorno, saindo do Centro, sem ter vindo do
    # bairro (tipo "baixada", IDA vazia). O horario delas E uma passagem pelo
    # ponto, mas o onibus PARTE dali — a UI precisa marcar isso.
    passagens_cortadas = sorted(
        v["volta"] for t in turnos for tb in t["tabelas"] for v in tb["viagens"]
        if v["tipo"] == "baixada" and v["volta"])

    # Painel de frequencia: partidas do terminal de ORIGEM. Mantido como
    # referencia e para validar a integridade da planilha. NAO e o que o
    # site mostra ao operador.
    saidas_origem = [f["hora"] for f in freq]
    # Uma partida do painel pode corresponder ao IDA de uma viagem normal ou
    # ao VOLTA — nas baixadas (viagem que começa no retorno) e, na 305, nas
    # partidas do outro terminal, que o painel também conta.
    partidas = set()
    for t in turnos:
        for tb in t["tabelas"]:
            for v in tb["viagens"]:
                if v["continuacao"]:
                    continue
                partidas.update(x for x in (v["ida"], v["volta"]) if x)
    conflitos = []
    so_painel = [h for h in saidas_origem if h not in partidas]
    so_quadro = sorted(
        v["ida"] for t in turnos for tb in t["tabelas"] for v in tb["viagens"]
        if v["tipo"] == "normal" and v["ida"] not in set(saidas_origem))
    if so_painel or so_quadro:
        conflitos.append({
            "tipo": "painel-x-quadro",
            "soNoPainel": so_painel,
            "soNoQuadro": so_quadro,
            "nota": "Divergência na planilha de origem. As saídas publicadas seguem "
                    "o painel de frequência, que bate com o total declarado.",
        })

    return {
        "tipoDia": tipo_dia,
        "totalViagens": total,
        "passagens": passagens,
        "passagensCortadas": passagens_cortadas,
        "saidasOrigem": saidas_origem,
        "frequencia": freq,
        "origemMista": any(f["origem"] for f in freq),
        "turnos": turnos,
        "conflitos": conflitos,
    }, empresa


def ler_aba(ws, aba):
    linhas = list(ws.iter_rows(values_only=True))
    blocos = achar_blocos_de_dia(linhas)

    if not blocos:
        # Aba de escala regular: um unico bloco, sem cabecalho de dia do evento.
        # A estrutura interna e identica a das abas do evento (turnos, TABs,
        # painel de frequencia, "N VIAGENS"), entao passa pelo mesmo leitor e o
        # resultado e replicado nos 3 dias — 05 e 07/09 sao feriados e a linha
        # roda a mesma escala nos tres.
        base = ler_escala_base(linhas)
        bloco, empresa = ler_bloco(linhas, 0, len(linhas), len(linhas), aba,
                                   "escala regular")
        observacoes, legenda = ler_notas(linhas)
        dias = {dia: copy.deepcopy(bloco) for dia in DIAS}
        return dias, empresa, observacoes, legenda, base

    if len(blocos) != 3:
        avisar("%s: %d blocos de dia do evento (esperado 3)" % (aba, len(blocos)))

    # Cabeçalhos de painel de frequência: servem de limite superior para a
    # leitura do painel do dia anterior, que pode transbordar o bloco.
    cabs_freq = [i for i, r in enumerate(linhas)
                 if any(txt(c) == "HORA" for c in r)]

    dias = {}
    empresa = None
    for k, (ini, fim) in enumerate(blocos[:3]):
        dia = DIAS[k]
        bloco = linhas[ini:fim]
        limite = next((i for i in cabs_freq if i >= fim), len(linhas))

        # Confere o rótulo da planilha contra a posição. A aba 448 tem o 3º bloco
        # rotulado "05/09/2026" por erro de digitação da origem: a posição vence.
        for c in bloco[0][:4]:
            rotulo = None
            if isinstance(c, datetime.datetime) and c.year == 2026:
                rotulo = c.strftime("%Y-%m-%d")
            elif isinstance(c, str):
                m = re.match(r"\s*(0[567])/09/2026", c)
                if m:
                    rotulo = "2026-09-" + m.group(1)
            if rotulo and rotulo != dia:
                avisar("%s: bloco %d rotulado %s na planilha; usando %s (posição)"
                       % (aba, k + 1, rotulo, dia))
            if rotulo:
                break

        dias[dia], emp = ler_bloco(linhas, ini, fim, limite, aba, dia)
        empresa = empresa or emp

    observacoes, legenda = ler_notas(linhas)
    return dias, empresa, observacoes, legenda, None


# ------------------------------------------------------------------ validações

def validar(linhas_json):
    erros = []
    for num, L in sorted(linhas_json.items()):
        # Vale para TODA linha publicada com quadro — as de evento e as de
        # escala regular, que desde a §1.2 tambem entram em dias{}.
        if not L["dias"]:
            erros.append("%s: publicada sem nenhum quadro horário" % num)
            continue
        for dia in DIAS:
            b = L["dias"].get(dia)
            if not b:
                erros.append("%s/%s: dia ausente" % (num, dia))
                continue
            s = b["saidasOrigem"]
            pa = b["passagens"]
            rev = [v for t in b["turnos"] for tb in t["tabelas"]
                   for v in tb["viagens"] if v["tipo"] != "continuacao"]
            if not pa:
                erros.append("%s/%s: sem passagens pelo ponto" % (num, dia))
            if len(pa) != len(rev):
                erros.append("%s/%s: %d passagens != %d viagens de receita "
                             "(viagem sem VOLTA)" % (num, dia, len(pa), len(rev)))
            if pa != sorted(pa):
                erros.append("%s/%s: passagens fora de ordem" % (num, dia))
            if not s:
                erros.append("%s/%s: sem saídas de origem" % (num, dia))
            if b["totalViagens"] is None:
                erros.append("%s/%s: sem 'N VIAGENS'" % (num, dia))
            elif len(s) != b["totalViagens"]:
                erros.append("%s/%s: %d saídas != %d VIAGENS declaradas"
                             % (num, dia, len(s), b["totalViagens"]))
            if s != sorted(s):
                erros.append("%s/%s: saídas fora de ordem cronológica" % (num, dia))
            for h in s:
                if not re.match(r"^\d{2}:\d{2}$", h) or not (0 <= int(h[:2]) <= 29):
                    erros.append("%s/%s: horário inválido %r" % (num, dia, h))
            # O painel e o quadro devem descrever as mesmas partidas.
            for c in b["conflitos"]:
                erros.append(
                    "%s/%s: CONFLITO painel x quadro — só no painel: [%s] | só no quadro: [%s]"
                    % (num, dia, ", ".join(c["soNoPainel"]) or "-",
                       ", ".join(c["soNoQuadro"]) or "-"))
            for t in b["turnos"]:
                for tb in t["tabelas"]:
                    j = tb["jornada"]
                    if j and int(j[:2]) < 2:
                        erros.append("%s/%s/%s: jornada suspeita %s (coluna errada?)"
                                     % (num, dia, tb["tab"], j))
    return erros


# ----------------------------------------------------------------------- saída

def main():
    linhas_json = {}
    ponto_de = {n: p for p, ns in PONTO_POR_LINHA.items() for n in ns}
    nomes = ler_nomes()

    for g in GRUPOS:
        wb = openpyxl.load_workbook(PLANILHAS / g, data_only=True)
        for aba in wb.sheetnames:
            num = str(aba).strip()
            if num in linhas_json:
                anterior = linhas_json[num]["grupoOrigem"]
                preferido = GRUPO_DO_PONTO.get(ponto_de.get(num), anterior)
                avisar("%s: aba duplicada em %s e %s; usando %s (grupo do ponto %s)"
                       % (num, anterior, g, preferido, ponto_de.get(num)))
                if preferido != g:
                    continue
            dias, empresa, observacoes, legenda, base = ler_aba(wb[aba], num)
            if num not in ponto_de:
                avisar("%s: linha sem ponto mapeado (arquivo %s)" % (num, g))
            linhas_json[num] = {
                "numero": num,
                "nome": nomes.get(num),
                "empresa": empresa,
                "grupoOrigem": g,
                "pontoId": ponto_de.get(num),
                "temProgramacaoEvento": base is None,
                "escalaBase": base,
                "avisos": ([a["id"] for a in AVISOS
                            if a["linhas"] == TODAS or num in a["linhas"]]
                           if base is None else []),
                "observacoes": observacoes,
                "legendaViagens": legenda,
                "dias": dias or {},
            }
            if num not in nomes:
                avisar("%s: linha ausente de nomes_linhas.json; publicada sem "
                       "itinerário" % num)
            if base is not None:
                avisar("%s: aba e escala regular (ref. %s, %s), sem programacao do "
                       "evento; quadro replicado nos 3 dias"
                       % (num, base["dataReferencia"], base["tipoDia"]))

    for p in PONTOS:
        no_ponto = [n for n in linhas_json if linhas_json[n]["pontoId"] == p["id"]]
        p["linhas"] = sorted(no_ponto)
        p["linhasComProgramacao"] = sorted(
            n for n in no_ponto if linhas_json[n]["temProgramacaoEvento"])
        p["linhasSemProgramacao"] = sorted(
            n for n in no_ponto if not linhas_json[n]["temProgramacaoEvento"])
        p["linhasSemPlanilha"] = sorted(n for n in PONTO_POR_LINHA[p["id"]]
                                        if n not in linhas_json)
        if p["linhasSemPlanilha"]:
            avisar("%s: linhas da arte sem aba nenhuma: %s"
                   % (p["id"], ", ".join(p["linhasSemPlanilha"])))

    doc = {
        "versaoDados": datetime.datetime.now().astimezone().replace(microsecond=0).isoformat(),
        "evento": {"nome": "Sou Manaus Passo a Paço 2026", "dias": DIAS},
        "pontos": PONTOS,
        "linhas": OrderedDict(sorted(linhas_json.items())),
        "avisos": AVISOS,
    }

    erros = validar(linhas_json)

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    with open(SAIDA, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=1)

    total_viagens = sum(len(tb["viagens"])
                        for L in linhas_json.values() for b in L["dias"].values()
                        for t in b["turnos"] for tb in t["tabelas"])
    total_saidas = sum(len(b["saidasOrigem"]) for L in linhas_json.values()
                       for b in L["dias"].values())
    total_passagens = sum(len(b["passagens"]) for L in linhas_json.values()
                          for b in L["dias"].values())

    com = [n for n, L in linhas_json.items() if L["temProgramacaoEvento"]]
    sem = [n for n, L in linhas_json.items() if not L["temProgramacaoEvento"]]

    rel = ["RELATÓRIO DE CONVERSÃO — %s" % doc["versaoDados"], ""]
    rel.append("Linhas na base: %d" % len(linhas_json))
    rel.append("  com programação do evento: %d (%d blocos linha-dia)"
               % (len(com), len(com) * 3))
    rel.append("  só escala regular, sem evento: %d -> %s"
               % (len(sem), ", ".join(sorted(sem))))
    rel.append("Passagens pelo ponto (VOLTA, publicadas): %d" % total_passagens)
    rel.append("Saídas do terminal de origem (IDA, referência): %d" % total_saidas)
    rel.append("Viagens no quadro completo: %d" % total_viagens)
    rel.append("")
    rel.append("%-6s %-12s %5s %5s %5s %5s %-8s %-8s" %
               ("LINHA", "DIA", "DECL", "ORIG", "PASSA", "VIAG",
                "1ªPASSAG", "ÚLTIMA"))
    for num, L in sorted(linhas_json.items()):
        if not L["temProgramacaoEvento"]:
            continue
        for dia in DIAS:
            b = L["dias"].get(dia)
            if not b:
                rel.append("%-6s %-12s  --- AUSENTE ---" % (num, dia))
                continue
            s = b["saidasOrigem"]
            pa = b["passagens"]
            nv = sum(len(tb["viagens"]) for t in b["turnos"] for tb in t["tabelas"])
            rel.append("%-6s %-12s %5s %5d %5d %5d %-8s %-8s" %
                       (num, dia, b["totalViagens"], len(s), len(pa), nv,
                        pa[0] if pa else "-", pa[-1] if pa else "-"))
    rel.append("")
    rel.append("AVISOS DE CONVERSÃO (%d):" % len(avisos_conversao))
    rel += ["  " + a for a in avisos_conversao] or ["  nenhum"]
    rel.append("")
    rel.append("ERROS DE VALIDAÇÃO (%d):" % len(erros))
    rel += ["  " + e for e in erros] or ["  nenhum"]

    RELATORIO.parent.mkdir(parents=True, exist_ok=True)
    RELATORIO.write_text("\n".join(rel), encoding="utf-8")

    print("\n".join(rel[:6]))
    print("\nAvisos: %d | Erros: %d" % (len(avisos_conversao), len(erros)))
    print("JSON: %s" % SAIDA)
    print("Relatório: %s" % RELATORIO)
    return 1 if erros else 0


if __name__ == "__main__":
    sys.exit(main())
